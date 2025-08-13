"""
Class Routes Module for Offline Attendance System

This module contains all class table management endpoints:
- Class record upload from Excel files
- Class table management and operations
- Class data retrieval and display
- Class table deletion and cleanup

Class Management Features:
- Upload class records from Excel files with metadata extraction
- Create and manage class-specific tables
- Display class information with proper formatting
- Delete class tables and associated data
"""

import sqlite3
import openpyxl
import csv
from io import StringIO
import re
from flask import Blueprint, request, jsonify
from config.config import Config
from database.class_table_manager import (
    create_class_table, insert_students as insert_class_students,
    OptimizedClassManager, create_class_optimized
)

class_bp = Blueprint('class', __name__)

def convert_year_to_integer(year_level):
    """Convert year level to integer for database storage"""
    if isinstance(year_level, int):
        return year_level
    
    year_str = str(year_level).lower().strip()
    
    # Extract numeric value from common year level formats
    if '1st' in year_str or 'first' in year_str or year_str == '1':
        return 1
    elif '2nd' in year_str or 'second' in year_str or year_str == '2':
        return 2
    elif '3rd' in year_str or 'third' in year_str or year_str == '3':
        return 3
    elif '4th' in year_str or 'fourth' in year_str or year_str == '4':
        return 4
    elif '5th' in year_str or 'fifth' in year_str or year_str == '5':
        return 5
    
    # Try to extract any number from the string
    import re
    match = re.search(r'(\d+)', year_str)
    if match:
        return int(match.group(1))
    
    # Default to 1 if no valid year found
    print(f"Warning: Unable to parse year level: {year_level}, defaulting to 1")
    return 1

@class_bp.route('/upload_class_record', methods=['POST'])
def upload_class_record():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file part in the request'}), 400
            
        file = request.files['file']
        file_name = file.filename  # <-- Move this up before using it
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
            
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Only Excel files (.xlsx, .xls) and CSV files (.csv) are allowed'}), 400

        filename_lower = file.filename.lower()
        
        # Process based on file type
        if filename_lower.endswith(('.xlsx', '.xls')):
            # Excel file processing
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 5:  # At least metadata + headers + 1 student
                return jsonify({'error': 'Invalid file format - not enough rows'}), 400

            # Extract metadata (first few rows)
            metadata = {}
            
            for i, row in enumerate(rows[:8]):  # Check first 8 rows for metadata
                if not any(row):  # Skip empty rows
                    continue
                    
                # Look for specific metadata patterns
                if row[0] and len(row) > 1:
                    key = str(row[0]).lower().strip()
                    value = str(row[1]).strip() if row[1] else None
                    
                    if 'professor' in key and value:
                        metadata['professor'] = value
                    elif 'class' in key and 'name' in key and value:
                        metadata['class_name'] = value
                    elif 'room' in key and 'type' in key and value:
                        metadata['room_type'] = value
                    elif 'building' in key and value:
                        metadata['building'] = value
                    elif 'venue' in key and value:
                        metadata['venue'] = value

            # Find the row with student headers
            student_data_start = None
            for i, row in enumerate(rows):
                if row and row[0] and 'student id' in str(row[0]).lower():
                    student_data_start = i
                    break
                    
            if student_data_start is None:
                return jsonify({'error': 'Could not find student data headers'}), 400

            # Process headers
            headers = [str(h).strip().lower().replace(' ', '_') for h in rows[student_data_start]]
            required_columns = {'student_id', 'student_name', 'year_level', 'course'}
            
            # Validate headers
            missing_columns = required_columns - set(headers)
            if missing_columns:
                return jsonify({
                    'error': f'Missing required columns: {", ".join(missing_columns)}. Found columns: {", ".join(headers)}'
                }), 400

            # Process student data (rows after headers)
            student_data = []

            for row in rows[student_data_start+1:]:
                # Skip empty rows or rows with empty student_id
                if not row or not row[0] or not str(row[0]).strip():
                    continue
                    
                student = dict(zip(headers, row))
                student_id = str(student.get('student_id', '')).strip()
                
                student_data.append({
                    'studentId': student_id,
                    'studentName': str(student.get('student_name', '')).strip(),
                    'yearLevel': str(student.get('year_level', '')).strip(),
                    'course': str(student.get('course', '')).strip()
                })

        elif filename_lower.endswith('.csv'):
            # CSV file processing
            try:
                # Try different encodings to handle various CSV file formats
                raw_content = file.read()
                
                # Try UTF-8 first (most common)
                try:
                    content = raw_content.decode('utf-8')
                except UnicodeDecodeError:
                    # Try UTF-8 with BOM
                    try:
                        content = raw_content.decode('utf-8-sig')
                    except UnicodeDecodeError:
                        # Try Windows-1252 (common for Excel-generated CSV)
                        try:
                            content = raw_content.decode('windows-1252')
                        except UnicodeDecodeError:
                            # Try ISO-8859-1 / Latin-1 (fallback)
                            try:
                                content = raw_content.decode('latin-1')
                            except UnicodeDecodeError:
                                # Last resort: CP1252
                                try:
                                    content = raw_content.decode('cp1252')
                                except UnicodeDecodeError:
                                    # If all fail, try to decode with error handling
                                    content = raw_content.decode('utf-8', errors='replace')
                                    
            except Exception as e:
                return jsonify({'error': f'Error reading CSV file: {str(e)}'}), 400

            # Parse CSV content
            reader = csv.reader(StringIO(content))
            rows = list(reader)
            
            if len(rows) < 2:  # At least headers + 1 student row
                return jsonify({'error': 'CSV must have at least header and one data row'}), 400

            # For CSV, we'll look for metadata in the first few rows, similar to Excel
            metadata = {}
            student_data_start = None
            
            # Look for metadata patterns in first few rows
            for i, row in enumerate(rows[:10]):  # Check more rows
                if not row or not any(str(cell).strip() for cell in row if cell is not None):  # Skip empty rows
                    continue
                
                # Check for the specific format from your Excel (metadata keys in first row, values in second)
                if i < len(rows) - 1:  # Ensure there's a next row
                    next_row = rows[i + 1] if i + 1 < len(rows) else []
                    
                    # Check if current row contains metadata headers and next row has values
                    if row and len(row) >= 2 and next_row and len(next_row) >= 2:
                        current_row_text = ' '.join(str(cell).lower() for cell in row if cell).strip()
                        
                        # Look for metadata pattern: "professor", "room type", "venue" etc in current row
                        if any(keyword in current_row_text for keyword in ['professor', 'instructor', 'teacher', 'room', 'venue', 'building']):
                            # Extract metadata from paired rows
                            for j, header_cell in enumerate(row):
                                if j < len(next_row) and header_cell and next_row[j]:
                                    key = str(header_cell).lower().strip()
                                    value = str(next_row[j]).strip()
                                    
                                    if any(word in key for word in ['professor', 'instructor', 'teacher']):
                                        metadata['professor'] = value
                                    elif any(word in key for word in ['room', 'venue', 'building']):
                                        if 'type' in key:
                                            metadata['room_type'] = value
                                        elif 'building' in key or 'venue' in key:
                                            metadata['venue'] = value
                                            metadata['building'] = value
                                    elif any(word in key for word in ['class', 'course', 'subject']):
                                        metadata['class_name'] = value
                    
                # Handle single column metadata (like "Professor: John Doe")
                if len(row) >= 1 and row[0]:
                    first_cell = str(row[0]).strip()
                    
                    # Check for colon-separated metadata in single cell
                    if ':' in first_cell:
                        parts = first_cell.split(':', 1)
                        if len(parts) == 2:
                            key = parts[0].lower().strip()
                            value = parts[1].strip()
                            
                            # More flexible professor name detection
                            if any(word in key for word in ['professor', 'instructor', 'teacher', 'prof']):
                                if value:
                                    metadata['professor'] = value
                            elif any(word in key for word in ['class', 'course', 'subject']):
                                if 'name' in key and value:
                                    metadata['class_name'] = value
                            elif 'room' in key and value:
                                metadata['room_type'] = value
                            elif 'building' in key and value:
                                metadata['building'] = value
                            elif 'venue' in key and value:
                                metadata['venue'] = value
                
                # Handle two-column metadata (key in col 1, value in col 2)
                if len(row) >= 2 and row[0] and row[1]:
                    key = str(row[0]).lower().strip()
                    value = str(row[1]).strip()
                    
                    # More flexible professor name detection
                    if any(word in key for word in ['professor', 'instructor', 'teacher', 'prof']):
                        if value:
                            metadata['professor'] = value
                    elif any(word in key for word in ['class', 'course', 'subject']):
                        if 'name' in key and value:
                            metadata['class_name'] = value
                        elif not 'name' in key and value:  # Just "class" or "course"
                            metadata['class_name'] = value
                    elif 'room' in key and 'type' in key and value:
                        metadata['room_type'] = value
                    elif 'building' in key and value:
                        metadata['building'] = value
                    elif 'venue' in key and value:
                        metadata['venue'] = value
                
                # Check if this row contains student headers (but don't break yet, continue looking for metadata)
                if row and len(row) > 0:
                    row_text = ' '.join(str(cell).lower() for cell in row if cell).strip()
                    if any(pattern in row_text for pattern in ['student id', 'student_id', 'studentid']):
                        if student_data_start is None:  # Only set if not already found
                            student_data_start = i

            # If no explicit student data start found, look for headers more broadly
            if student_data_start is None:
                for i, row in enumerate(rows):
                    if row and len(row) > 2:  # Must have at least 3 columns for student data
                        row_text = ' '.join(str(cell).lower() for cell in row if cell).strip()
                        # Look for typical student data headers
                        if any(pattern in row_text for pattern in ['name', 'course', 'year', 'id']):
                            student_data_start = i
                            break
                
                # Last resort: assume first non-empty row is headers
                if student_data_start is None:
                    for i, row in enumerate(rows):
                        if row and any(str(cell).strip() for cell in row if cell is not None):
                            student_data_start = i
                            break

            # Normalize headers for matching
            def normalize_header(header):
                return str(header).lower().strip().replace(' ', '_').replace('-', '_')

            headers = [normalize_header(h) for h in rows[student_data_start]]
            
            # Build header mapping
            header_map = {}
            for idx, header in enumerate(headers):
                header_map[header] = idx
            
            # Check for required columns (flexible matching)
            required_mappings = {
                'student_id': ['student_id', 'studentid', 'id', 'student_number', 'studentnumber'],
                'student_name': ['student_name', 'studentname', 'name', 'student', 'full_name', 'fullname'],
                'year_level': ['year_level', 'yearlevel', 'year', 'level', 'grade'],
                'course': ['course', 'program', 'major', 'subject']
            }
            
            column_indices = {}
            missing_columns = []
            
            for required_key, possible_headers in required_mappings.items():
                found = False
                for possible in possible_headers:
                    if possible in header_map:
                        column_indices[required_key] = header_map[possible]
                        found = True
                        break
                if not found:
                    missing_columns.append(required_key)
            
            if missing_columns:
                return jsonify({
                    'error': f'Missing required columns: {", ".join(missing_columns)}. Found columns: {", ".join([rows[student_data_start][i] for i in range(len(rows[student_data_start]))])}'
                }), 400

            # Process student data
            student_data = []
            for row in rows[student_data_start+1:]:
                # Skip empty rows
                if not row or not any(cell.strip() for cell in row if cell):
                    continue
                
                # Extract values using column indices
                try:
                    student_id = str(row[column_indices['student_id']]).strip() if column_indices['student_id'] < len(row) else ''
                    if not student_id:
                        continue
                        
                    student_name = str(row[column_indices['student_name']]).strip() if column_indices['student_name'] < len(row) else ''
                    year_level = str(row[column_indices['year_level']]).strip() if column_indices['year_level'] < len(row) else ''
                    course = str(row[column_indices['course']]).strip() if column_indices['course'] < len(row) else ''
                    
                    student_data.append({
                        'studentId': student_id,
                        'studentName': student_name,
                        'yearLevel': year_level,
                        'course': course
                    })
                except IndexError:
                    continue  # Skip rows with missing data

            # Set defaults if metadata not found
            metadata = {}
        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        # Set defaults if not found
        professor_name = metadata.get('professor', 'Unknown Professor')
        class_name = metadata.get('class_name', file_name)  # Use extracted class name or file name
        room_type = metadata.get('room_type', 'Classroom')
        venue = metadata.get('venue', 'Main Building')
        building = metadata.get('building', 'Main Building')

        if not student_data:
            return jsonify({'error': 'No valid student data found in the file'}), 400

        # Create display name from filename (without extension) and professor name
        if file_name.lower().endswith('.xlsx'):
            file_name = file_name[:-5]
        elif file_name.lower().endswith('.xls'):
            file_name = file_name[:-4]
            
        # Ensure there's exactly one " - " between filename and professor name
        file_name = file_name.rstrip(' -')  # Remove any existing dashes or spaces at the end
        professor_name = professor_name.lstrip(' -')  # Remove any existing dashes or spaces at the start
        display_name = f"{file_name} - {professor_name}"

        # Create table name (sanitized version for database)
        # Ensure there's exactly one "_" between filename and professor name in table name
        sanitized_file_name = file_name.replace(' ', '_').rstrip('_')
        sanitized_professor = professor_name.replace(' ', '_').lstrip('_')
        table_name = f"{sanitized_file_name}___{sanitized_professor}"
        table_name = ''.join(c for c in table_name if c.isalnum() or c == '_')

        # Database operations - Choose between old redundant method and new optimized method
        use_optimized = request.form.get('use_optimized', 'false').lower() == 'true'
        
        if use_optimized:
            # NEW OPTIMIZED METHOD - eliminates data redundancy
            print("Using optimized classes database schema...")
            manager = OptimizedClassManager()
            
            class_id = manager.import_from_excel_data(
                class_name=class_name,  # Use extracted class name
                professor_name=professor_name,
                student_data=student_data
            )
            
            if class_id:
                success_message = f'Successfully imported {len(student_data)} students using optimized schema'
            else:
                return jsonify({'error': 'Failed to create class with optimized schema'}), 500
        else:
            # OLD METHOD - creates redundant table per class (for backward compatibility)
            print("Using legacy table-per-class method...")
            columns = [
                ('student_id', 'TEXT'),
                ('student_name', 'TEXT'),
                ('year_level', 'TEXT'),
                ('course', 'TEXT')
            ]
            
            create_class_table(table_name, columns, db_path=Config.CLASSES_DATABASE_PATH)
            insert_class_students(table_name, student_data, db_path=Config.CLASSES_DATABASE_PATH)
            success_message = f'Successfully imported {len(student_data)} students to class table'
        
        # Skip attendance.db insertion for uploaded class records
        # Class records are managed separately from the main attendance system
        print("Skipping attendance.db insertion for uploaded class records - class tables are managed separately")

        return jsonify({
            'message': success_message,
            'display_name': display_name,
            'professor': professor_name,
            'room_type': room_type,
            'venue': venue,
            'optimized': use_optimized,
            'student_data': student_data
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': f'Server error: {str(e)}'
        }), 500

@class_bp.route('/preview_class_record', methods=['POST'])
def preview_class_record():
    """Preview class record data before saving"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file part in the request'}), 400
            
        file = request.files['file']
        file_name = file.filename
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
            
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            return jsonify({'error': 'Only Excel files (.xlsx, .xls) and CSV files (.csv) are allowed'}), 400

        filename_lower = file.filename.lower()
        
        # Use the same parsing logic as upload_class_record but don't save to database
        student_data = []
        metadata = {}
        
        # Process based on file type (same logic as upload_class_record)
        if filename_lower.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 5:
                return jsonify({'error': 'Invalid file format - not enough rows'}), 400

            # Extract metadata (first few rows)
            for i, row in enumerate(rows[:8]):
                if not any(row):
                    continue
                    
                if row[0] and len(row) > 1:
                    key = str(row[0]).lower().strip()
                    value = str(row[1]).strip() if row[1] else None
                    
                    if 'professor' in key and value:
                        metadata['professor'] = value
                    elif 'class' in key and 'name' in key and value:
                        metadata['class_name'] = value
                    elif 'room' in key and 'type' in key and value:
                        metadata['room_type'] = value
                    elif 'building' in key and value:
                        metadata['building'] = value
                    elif 'venue' in key and value:
                        metadata['venue'] = value

            # Find student data headers
            student_data_start = None
            for i, row in enumerate(rows):
                if row and row[0] and 'student id' in str(row[0]).lower():
                    student_data_start = i
                    break
                    
            if student_data_start is None:
                return jsonify({'error': 'Could not find student data headers'}), 400

            headers = [str(h).strip().lower().replace(' ', '_') for h in rows[student_data_start]]
            required_columns = {'student_id', 'student_name', 'year_level', 'course'}
            
            missing_columns = required_columns - set(headers)
            if missing_columns:
                return jsonify({
                    'error': f'Missing required columns: {", ".join(missing_columns)}'
                }), 400

            # Process student data
            for row in rows[student_data_start+1:]:
                if not row or not row[0] or not str(row[0]).strip():
                    continue
                    
                student = dict(zip(headers, row))
                student_id = str(student.get('student_id', '')).strip()
                
                student_data.append({
                    'studentId': student_id,
                    'studentName': str(student.get('student_name', '')).strip(),
                    'yearLevel': str(student.get('year_level', '')).strip(),
                    'course': str(student.get('course', '')).strip()
                })

        elif filename_lower.endswith('.csv'):
            # CSV processing (similar to upload_class_record)
            try:
                raw_content = file.read()
                
                try:
                    content = raw_content.decode('utf-8')
                except UnicodeDecodeError:
                    try:
                        content = raw_content.decode('utf-8-sig')
                    except UnicodeDecodeError:
                        try:
                            content = raw_content.decode('windows-1252')
                        except UnicodeDecodeError:
                            try:
                                content = raw_content.decode('latin-1')
                            except UnicodeDecodeError:
                                content = raw_content.decode('utf-8', errors='replace')
                                    
            except Exception as e:
                return jsonify({'error': f'Error reading CSV file: {str(e)}'}), 400

            import csv
            reader = csv.reader(StringIO(content))
            rows = list(reader)
            
            if len(rows) < 2:
                return jsonify({'error': 'CSV must have at least header and one data row'}), 400

            # Extract metadata from CSV (similar logic)
            student_data_start = None
            for i, row in enumerate(rows[:10]):
                if not row or not any(str(cell).strip() for cell in row if cell is not None):
                    continue
                
                # Check for metadata patterns
                if len(row) >= 2 and row[0] and row[1]:
                    key = str(row[0]).lower().strip()
                    value = str(row[1]).strip()
                    
                    if any(word in key for word in ['professor', 'instructor', 'teacher', 'prof']):
                        if value:
                            metadata['professor'] = value
                
                # Check for student headers
                if row and len(row) > 0:
                    row_text = ' '.join(str(cell).lower() for cell in row if cell).strip()
                    if any(pattern in row_text for pattern in ['student id', 'student_id', 'studentid']):
                        if student_data_start is None:
                            student_data_start = i

            if student_data_start is None:
                for i, row in enumerate(rows):
                    if row and len(row) > 2:
                        row_text = ' '.join(str(cell).lower() for cell in row if cell).strip()
                        if any(pattern in row_text for pattern in ['name', 'course', 'year', 'id']):
                            student_data_start = i
                            break

            def normalize_header(header):
                return str(header).lower().strip().replace(' ', '_').replace('-', '_')

            headers = [normalize_header(h) for h in rows[student_data_start]]
            
            # Build header mapping
            header_map = {}
            for idx, header in enumerate(headers):
                header_map[header] = idx
            
            required_mappings = {
                'student_id': ['student_id', 'studentid', 'id', 'student_number', 'studentnumber'],
                'student_name': ['student_name', 'studentname', 'name', 'student', 'full_name', 'fullname'],
                'year_level': ['year_level', 'yearlevel', 'year', 'level', 'grade'],
                'course': ['course', 'program', 'major', 'subject']
            }
            
            column_indices = {}
            missing_columns = []
            
            for required_key, possible_headers in required_mappings.items():
                found = False
                for possible in possible_headers:
                    if possible in header_map:
                        column_indices[required_key] = header_map[possible]
                        found = True
                        break
                if not found:
                    missing_columns.append(required_key)
            
            if missing_columns:
                return jsonify({
                    'error': f'Missing required columns: {", ".join(missing_columns)}'
                }), 400

            # Process student data
            for row in rows[student_data_start+1:]:
                if not row or not any(cell.strip() for cell in row if cell):
                    continue
                
                try:
                    student_id = str(row[column_indices['student_id']]).strip() if column_indices['student_id'] < len(row) else ''
                    if not student_id:
                        continue
                        
                    student_name = str(row[column_indices['student_name']]).strip() if column_indices['student_name'] < len(row) else ''
                    year_level = str(row[column_indices['year_level']]).strip() if column_indices['year_level'] < len(row) else ''
                    course = str(row[column_indices['course']]).strip() if column_indices['course'] < len(row) else ''
                    
                    student_data.append({
                        'studentId': student_id,
                        'studentName': student_name,
                        'yearLevel': year_level,
                        'course': course
                    })
                except IndexError:
                    continue

        # Set defaults if not found
        professor_name = metadata.get('professor', 'Unknown Professor')
        class_name = metadata.get('class_name', file_name)
        room_type = metadata.get('room_type', 'Classroom')
        venue = metadata.get('venue', 'Main Building')
        building = metadata.get('building', 'Main Building')

        if not student_data:
            return jsonify({'error': 'No valid student data found in the file'}), 400

        # Create display name
        if file_name.lower().endswith('.xlsx'):
            file_name = file_name[:-5]
        elif file_name.lower().endswith('.xls'):
            file_name = file_name[:-4]
        elif file_name.lower().endswith('.csv'):
            file_name = file_name[:-4]
            
        file_name = file_name.rstrip(' -')
        professor_name = professor_name.lstrip(' -')
        display_name = f"{file_name} - {professor_name}"

        return jsonify({
            'status': 'preview',
            'message': 'File processed successfully - ready for review',
            'class_data': {
                'display_name': display_name,
                'class_name': class_name,
                'professor': professor_name,
                'room_type': room_type,
                'venue': venue,
                'building': building,
                'student_count': len(student_data)
            },
            'student_data': student_data
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': f'Server error: {str(e)}'
        }), 500

@class_bp.route('/save_class_record', methods=['POST'])
def save_class_record():
    """Save the reviewed class record data"""
    try:
        data = request.json or {}
        
        if not data.get('class_data') or not data.get('student_data'):
            return jsonify({'error': 'Missing class or student data'}), 400
        
        class_data = data['class_data']
        student_data = data['student_data']
        
        # Use optimized schema
        from database.class_table_manager import OptimizedClassManager
        manager = OptimizedClassManager()
        
        # Prepare metadata with room and venue info
        metadata = {
            'room_type': class_data.get('room_type'),
            'venue': class_data.get('venue'),
            'building': class_data.get('building')
        }
        
        class_id = manager.import_from_excel_data(
            class_name=class_data.get('class_name'),
            professor_name=class_data.get('professor'),
            student_data=student_data,
            metadata=metadata
        )
        
        success_message = f"Class '{class_data.get('display_name')}' created successfully with {len(student_data)} students!"
        
        return jsonify({
            'message': success_message,
            'class_id': class_id,
            'display_name': class_data.get('display_name'),
            'professor': class_data.get('professor'),
            'student_count': len(student_data)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': f'Server error: {str(e)}'
        }), 500

@class_bp.route('/api/class_tables', methods=['GET'])
def get_class_tables():
    try:
        db_path = Config.CLASSES_DATABASE_PATH
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Get all user tables
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
        tables = [row[0] for row in cursor.fetchall()]
        
        result = []
        for table in tables:
            # Get all students from each table
            cursor.execute(f'SELECT * FROM "{table}"')  # Note the quotes around table name
            columns = [desc[0] for desc in cursor.description]
            students = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Reconstruct display name from table name
            display_name = table.replace('_', ' ').replace('-', ' - ')
            
            result.append({
                'table_name': table,
                'display_name': display_name,
                'students': students,
                'can_delete': True  # Flag to indicate deletable tables
            })
            
        conn.close()
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to retrieve class tables: {str(e)}'
        }), 500

@class_bp.route('/api/delete_class_table', methods=['POST'])
def delete_class_table():
    try:
        data = request.json or {}
        table_name = data.get('table_name')
        
        if not table_name:
            return jsonify({'error': 'Table name is required'}), 400
            
        db_path = Config.CLASSES_DATABASE_PATH
        
        # Connect to the database
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Check if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Table not found'}), 404
            
        # Delete the table
        cursor.execute(f'DROP TABLE "{table_name}"')
        conn.commit()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'message': f'Table {table_name} deleted successfully'
        })
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to delete table: {str(e)}'
        }), 500

@class_bp.route('/api/classes')
def get_classes():
    db_path = Config.CLASSES_DATABASE_PATH
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    # Get all user tables (each table is a class)
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
    tables = [row[0] for row in cursor.fetchall()]
    conn.close()
    # Reconstruct display names (match your upload logic)
    class_list = []
    for table in tables:
        parts = table.split('___')
        if len(parts) == 2:
            file_part = parts[0].replace('_', ' ')
            prof_part = parts[1].replace('_', ' ')
            display_name = f"{file_part} - {prof_part}"
        else:
            display_name = table
        class_list.append({'table_name': table, 'display_name': display_name})
    return jsonify({'classes': class_list})

@class_bp.route('/api/classes/<table_name>/students', methods=['POST'])
def add_student_to_class(table_name):
    """Add a single student to a specific class table"""
    try:
        data = request.json or {}
        
        # Validate required fields
        required_fields = ['student_id', 'student_name', 'year_level', 'course']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # Sanitize input data
        student_data = {
            'studentId': str(data['student_id']).strip(),
            'studentName': str(data['student_name']).strip(),
            'yearLevel': str(data['year_level']).strip(),
            'course': str(data['course']).strip()
        }
        
        # Convert year_level to integer for attendance.db
        year_level_int = convert_year_to_integer(data['year_level'])
        
        # Validate table exists
        db_path = Config.CLASSES_DATABASE_PATH
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Check if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Class table not found'}), 404
        
        # Check if student already exists in this class
        cursor.execute(f'SELECT student_id FROM "{table_name}" WHERE student_id = ?', (student_data['studentId'],))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Student ID already exists in this class'}), 409
        
        # Insert the new student
        cursor.execute(f'''
            INSERT INTO "{table_name}" (student_id, student_name, year_level, course)
            VALUES (?, ?, ?, ?)
        ''', (student_data['studentId'], student_data['studentName'], 
              student_data['yearLevel'], student_data['course']))
        
        conn.commit()
        conn.close()
        
        # For manually added students, also add to attendance.db if not already there
        # (This is different from bulk uploads which are class-specific)
        try:
            attendance_conn = sqlite3.connect(Config.DATABASE_PATH)
            attendance_cursor = attendance_conn.cursor()
            
            # Check if student exists in attendance.db
            attendance_cursor.execute("SELECT student_id FROM students WHERE student_id = ?", (student_data['studentId'],))
            if not attendance_cursor.fetchone():
                # Add to attendance.db with integer year
                attendance_cursor.execute(
                    "INSERT INTO students (student_id, name, year, course) VALUES (?, ?, ?, ?)",
                    (student_data['studentId'], student_data['studentName'], 
                     year_level_int, student_data['course'])
                )
                attendance_conn.commit()
                print(f"Added manually entered student {student_data['studentId']} to attendance.db with year {year_level_int}")
            else:
                print(f"Student {student_data['studentId']} already exists in attendance.db")
            
            attendance_conn.close()
        except Exception as e:
            print(f"Warning: Could not add student to attendance.db: {str(e)}")
            # Continue even if attendance.db insertion fails
        
        return jsonify({
            'status': 'success',
            'message': 'Student added successfully',
            'student': student_data
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@class_bp.route('/api/classes/<table_name>/students/<student_id>', methods=['DELETE'])
def remove_student_from_class(table_name, student_id):
    """Remove a student from a specific class table"""
    try:
        # Validate table exists
        db_path = Config.CLASSES_DATABASE_PATH
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Check if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Class table not found'}), 404
        
        # Check if student exists in this class
        cursor.execute(f'SELECT student_name FROM "{table_name}" WHERE student_id = ?', (student_id,))
        student = cursor.fetchone()
        if not student:
            conn.close()
            return jsonify({'error': 'Student not found in this class'}), 404
        
        student_name = student[0]
        
        # Delete the student
        cursor.execute(f'DELETE FROM "{table_name}" WHERE student_id = ?', (student_id,))
        rows_affected = cursor.rowcount
        
        conn.commit()
        conn.close()
        
        if rows_affected > 0:
            return jsonify({
                'status': 'success',
                'message': f'Student {student_name} removed from class successfully'
            })
        else:
            return jsonify({'error': 'Failed to remove student'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

# ===================================================================
# OPTIMIZED CLASS API ENDPOINTS
# These endpoints use the new normalized schema instead of table-per-class
# ===================================================================

@class_bp.route('/api/optimized/setup', methods=['POST'])
def setup_optimized_schema():
    """Initialize the optimized classes database schema"""
    try:
        from database.models import create_optimized_classes_schema
        success = create_optimized_classes_schema()
        
        if success:
            return jsonify({
                'status': 'success',
                'message': 'Optimized classes schema created successfully'
            })
        else:
            return jsonify({'error': 'Failed to create optimized schema'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@class_bp.route('/api/optimized/migrate', methods=['POST'])
def migrate_to_optimized():
    """Migrate existing class tables to optimized schema"""
    try:
        from database.models import migrate_existing_classes_data
        success = migrate_existing_classes_data()
        
        if success:
            return jsonify({
                'status': 'success',
                'message': 'Successfully migrated to optimized schema'
            })
        else:
            return jsonify({'error': 'Migration failed'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@class_bp.route('/api/optimized/classes', methods=['GET'])
def get_optimized_classes():
    """Get all classes using the optimized schema"""
    try:
        manager = OptimizedClassManager()
        classes = manager.get_all_classes()
        
        return jsonify({
            'status': 'success',
            'classes': classes,
            'count': len(classes)
        })
        
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@class_bp.route('/api/optimized/classes/<int:class_id>/students', methods=['GET'])
def get_optimized_class_students(class_id):
    """Get students enrolled in a specific class using optimized schema"""
    try:
        manager = OptimizedClassManager()
        students = manager.get_class_students(class_id)
        
        return jsonify({
            'status': 'success',
            'students': students,
            'count': len(students),
            'class_id': class_id
        })
        
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@class_bp.route('/api/optimized/classes', methods=['POST'])
def create_optimized_class():
    """Create a new class using optimized schema"""
    try:
        data = request.get_json()
        
        if not data or not data.get('class_name') or not data.get('professor_name'):
            return jsonify({'error': 'class_name and professor_name are required'}), 400
        
        manager = OptimizedClassManager()
        class_id = manager.create_class(
            class_name=data['class_name'],
            professor_name=data['professor_name'],
            course_code=data.get('course_code'),
            semester=data.get('semester'),
            academic_year=data.get('academic_year')
        )
        
        # Note: room_type and venue are accepted but not yet stored in the optimized schema
        # These could be stored in a separate metadata table if needed in the future
        room_type = data.get('room_type', 'Classroom')
        venue = data.get('venue', 'Main Building')
        
        if class_id:
            return jsonify({
                'status': 'success',
                'message': f'Class created successfully in {room_type} at {venue}',
                'class_id': class_id,
                'room_type': room_type,
                'venue': venue
            })
        else:
            return jsonify({'error': 'Failed to create class'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@class_bp.route('/api/optimized/classes/<int:class_id>/enroll', methods=['POST'])
def enroll_students_optimized(class_id):
    """Enroll students in a class using optimized schema"""
    try:
        data = request.get_json()
        
        if not data or not data.get('student_ids'):
            return jsonify({'error': 'student_ids array is required'}), 400
        
        manager = OptimizedClassManager()
        enrolled_count = manager.enroll_students(class_id, data['student_ids'])
        
        return jsonify({
            'status': 'success',
            'message': f'Enrolled {enrolled_count} students successfully',
            'enrolled_count': enrolled_count,
            'total_requested': len(data['student_ids'])
        })
        
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@class_bp.route('/api/optimized/classes/<int:class_id>', methods=['DELETE'])
def delete_class_optimized(class_id):
    """Delete a class and all its enrollments using optimized schema"""
    try:
        manager = OptimizedClassManager()
        success = manager.delete_class(class_id)
        
        if success:
            return jsonify({
                'status': 'success',
                'message': f'Class {class_id} deleted successfully'
            })
        else:
            return jsonify({'error': 'Failed to delete class'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@class_bp.route('/api/optimized/classes/<int:class_id>/students/<student_id>', methods=['DELETE'])
def unenroll_student_optimized(class_id, student_id):
    """Remove a student from a class using optimized schema"""
    try:
        manager = OptimizedClassManager()
        success = manager.unenroll_student(class_id, student_id)
        
        if success:
            return jsonify({
                'status': 'success',
                'message': f'Student {student_id} removed from class successfully'
            })
        else:
            return jsonify({'error': 'Failed to remove student from class'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@class_bp.route('/api/classes/<int:class_id>/attendance-detail', methods=['GET'])
def get_class_attendance_detail(class_id):
    """Get detailed attendance data for a specific class showing student attendance across multiple sessions"""
    try:
        # Use the OptimizedClassManager to get students enrolled in this class
        manager = OptimizedClassManager()
        students = manager.get_class_students(class_id)
        
        if not students:
            return jsonify({
                'error': 'No students found in this class'
            }), 404
        
        # Connect to attendance database to get session data
        db_path = Config.DATABASE_PATH
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Get all attendance sessions for this class:
        # 1. Sessions where students from this class attended (via class_attendees)
        # 2. Sessions created specifically for this class (via class_table field)
        student_ids = [s['student_id'] for s in students]
        placeholders = ','.join(['?' for _ in student_ids])
        
        # Query for sessions where students attended OR sessions created for this class
        cursor.execute(f'''
            SELECT DISTINCT s.id, s.session_name, s.start_time, s.end_time, s.created_at
            FROM attendance_sessions s
            LEFT JOIN class_attendees ca ON s.id = ca.session_id
            WHERE (ca.student_id IN ({placeholders}) OR s.class_table = ?)
            ORDER BY s.created_at ASC
            LIMIT 50
        ''', student_ids + [str(class_id)])
        
        rows = cursor.fetchall()
        sessions = []
        for row in rows:
            sessions.append({
                'id': row[0],
                'session_name': row[1],
                'start_time': row[2],
                'end_time': row[3],
                'created_at': row[4]
            })
        
        # If no sessions found, return empty - don't show other classes' sessions
        if not sessions:
            conn.close()
            
            # Get class information
            classes_conn = sqlite3.connect(Config.CLASSES_DATABASE_PATH)
            classes_cursor = classes_conn.cursor()
            classes_cursor.execute('SELECT class_name, professor_name FROM classes WHERE id = ?', (class_id,))
            class_row = classes_cursor.fetchone()
            classes_conn.close()
            
            class_details = {
                'class_name': class_row[0] if class_row else 'Unknown',
                'professor_name': class_row[1] if class_row else 'Unknown'
            }
            
            return jsonify({
                'status': 'success',
                'class_id': class_id,
                'class_details': class_details,
                'students': {},
                'sessions': [],
                'total_students': len(students),
                'total_sessions': 0,
                'message': "No attendance sessions found for this class. Create a session and have students check in to see attendance data here."
            })
        
        # Get attendance data for each student in each session
        attendance_data = {}
        for student in students:
            student_id = student['student_id']
            attendance_data[student_id] = {
                'student_info': student,
                'sessions': {}
            }
            
            for session in sessions:
                session_id = session['id']
                
                # Check if student attended this session
                cursor.execute('''
                    SELECT ca.checked_in_at, ca.id
                    FROM class_attendees ca
                    WHERE ca.student_id = ? AND ca.session_id = ?
                ''', (student_id, session_id))
                
                attendance_record = cursor.fetchone()
                
                if attendance_record:
                    attendance_data[student_id]['sessions'][session_id] = {
                        'status': 'present',
                        'checked_in_at': attendance_record[0],
                        'attendance_id': attendance_record[1]
                    }
                else:
                    attendance_data[student_id]['sessions'][session_id] = {
                        'status': 'absent',
                        'checked_in_at': None,
                        'attendance_id': None
                    }
        
        conn.close()
        
        # Get class information
        classes_conn = sqlite3.connect(Config.CLASSES_DATABASE_PATH)
        classes_cursor = classes_conn.cursor()
        classes_cursor.execute('SELECT class_name, professor_name FROM classes WHERE id = ?', (class_id,))
        class_row = classes_cursor.fetchone()
        classes_conn.close()
        
        class_details = {
            'class_name': class_row[0] if class_row else 'Unknown',
            'professor_name': class_row[1] if class_row else 'Unknown'
        }
        
        # Create a message for empty sessions
        message = None
        if not sessions:
            message = "No attendance sessions found for this class. Create a session and have students check in to see attendance data here."
        elif len([s for s in sessions if 'note' not in s]) == 0:
            message = "Students from this class haven't attended any sessions yet. When they do, their attendance will appear here."
        
        return jsonify({
            'status': 'success',
            'class_id': class_id,
            'class_details': class_details,
            'students': attendance_data,
            'sessions': sessions,
            'total_students': len(students),
            'total_sessions': len(sessions),
            'message': message
        })
        
    except Exception as e:
        import traceback
        print("=== ERROR IN ATTENDANCE DETAIL ENDPOINT ===")
        traceback.print_exc()  # This will help us see the full error
        print("=== END ERROR ===")
        return jsonify({
            'error': f'Failed to retrieve attendance details: {str(e)}'
        }), 500

@class_bp.route('/api/classes/<table_name>/attendance-detail-legacy', methods=['GET'])
def get_class_attendance_detail_legacy(table_name):
    """Get detailed attendance data for a legacy class table"""
    try:
        db_path = Config.DATABASE_PATH
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Get all students from the legacy class table
        classes_conn = sqlite3.connect(Config.CLASSES_DATABASE_PATH)
        classes_cursor = classes_conn.cursor()
        
        # Check if table exists
        classes_cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not classes_cursor.fetchone():
            classes_conn.close()
            conn.close()
            return jsonify({'error': 'Class table not found'}), 404
        
        # Get students from legacy table
        classes_cursor.execute(f'SELECT student_id, student_name, year_level, course FROM "{table_name}"')
        students_data = [dict(row) for row in classes_cursor.fetchall()]
        classes_conn.close()
        
        if not students_data:
            conn.close()
            return jsonify({
                'error': 'No students found in this class'
            }), 404
        
        # Get all attendance sessions where these students participated
        student_ids = [s['student_id'] for s in students_data]
        placeholders = ','.join(['?' for _ in student_ids])
        
        cursor.execute(f'''
            SELECT DISTINCT s.id, s.session_name, s.start_time, s.end_time, s.created_at
            FROM attendance_sessions s
            INNER JOIN class_attendees ca ON s.id = ca.session_id
            WHERE ca.student_id IN ({placeholders})
            ORDER BY s.created_at DESC
        ''', student_ids)
        
        sessions = [dict(row) for row in cursor.fetchall()]
        
        # Get attendance data for each student in each session
        attendance_data = {}
        for student in students_data:
            student_id = student['student_id']
            attendance_data[student_id] = {
                'student_info': {
                    'student_id': student['student_id'],
                    'name': student['student_name'],
                    'year': student['year_level'],
                    'course': student['course']
                },
                'sessions': {}
            }
            
            for session in sessions:
                session_id = session['id']
                
                # Check if student attended this session
                cursor.execute('''
                    SELECT ca.checked_in_at, ca.id
                    FROM class_attendees ca
                    WHERE ca.student_id = ? AND ca.session_id = ?
                ''', (student_id, session_id))
                
                attendance_record = cursor.fetchone()
                
                if attendance_record:
                    attendance_data[student_id]['sessions'][session_id] = {
                        'status': 'present',
                        'checked_in_at': attendance_record[0],
                        'attendance_id': attendance_record[1]
                    }
                else:
                    attendance_data[student_id]['sessions'][session_id] = {
                        'status': 'absent',
                        'checked_in_at': None,
                        'attendance_id': None
                    }
        
        conn.close()
        
        return jsonify({
            'status': 'success',
            'table_name': table_name,
            'students': attendance_data,
            'sessions': sessions,
            'total_students': len(students_data),
            'total_sessions': len(sessions)
        })
        
    except Exception as e:
        return jsonify({
            'error': f'Failed to retrieve attendance details: {str(e)}'
        }), 500

@class_bp.route('/api/optimized/classes/<int:class_id>/unenroll', methods=['POST'])
def unenroll_students_optimized(class_id):
    """Remove multiple students from a class using optimized schema (bulk operation)"""
    try:
        print(f"DEBUG: unenroll_students_optimized called with class_id={class_id}")
        data = request.get_json()
        print(f"DEBUG: Request data: {data}")
        
        if not data or 'student_ids' not in data:
            print("DEBUG: Missing student_ids in request")
            return jsonify({'error': 'Missing student_ids in request'}), 400
        
        student_ids = data['student_ids']
        if not isinstance(student_ids, list) or not student_ids:
            print("DEBUG: student_ids must be a non-empty list")
            return jsonify({'error': 'student_ids must be a non-empty list'}), 400
        
        print(f"DEBUG: Processing student_ids: {student_ids}")
        manager = OptimizedClassManager()
        success_count = 0
        failed_students = []
        
        for student_id in student_ids:
            try:
                print(f"DEBUG: Attempting to unenroll student {student_id} from class {class_id}")
                success = manager.unenroll_student(class_id, student_id)
                if success:
                    success_count += 1
                    print(f"DEBUG: Successfully unenrolled {student_id}")
                else:
                    failed_students.append(student_id)
                    print(f"DEBUG: Failed to unenroll {student_id}")
            except Exception as e:
                print(f"DEBUG: Error unenrolling student {student_id}: {e}")
                failed_students.append(student_id)
        
        if success_count > 0:
            message = f'Successfully removed {success_count} student(s) from class'
            if failed_students:
                message += f'. Failed to remove: {", ".join(failed_students)}'
            
            print(f"DEBUG: Returning success response: {message}")
            return jsonify({
                'status': 'success',
                'message': message,
                'removed_count': success_count,
                'failed_students': failed_students
            })
        else:
            print("DEBUG: Failed to remove any students")
            return jsonify({'error': 'Failed to remove any students from class'}), 500
            
    except Exception as e:
        print(f"DEBUG: Exception in unenroll_students_optimized: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}'}), 500